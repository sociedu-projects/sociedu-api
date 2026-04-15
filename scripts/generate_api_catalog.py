# -*- coding: utf-8 -*-
"""Generate docs/API_Catalog_UniShare.xlsx from current API inventory."""
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
# Default output; override with env API_CATALOG_OUT if the xlsx is locked (e.g. open in Excel).
OUT = ROOT / "docs" / (os.environ.get("API_CATALOG_OUT") or "API_Catalog_UniShare.xlsx")

HEADERS = [
    "Module",
    "Tên API (VN)",
    "Method",
    "Endpoint",
    "JWT",
    "Kịch bản / mô tả",
    "Mã lỗi (errors.type)",
    "Kết quả",
    "HTTP",
]

JWT_KHONG = "Không"
JWT_CO = "Có"

# (module, name_vn, method, path, jwt, scenario, error_code, outcome, http)
ROWS = [
    # --- Auth ---
    ("Auth", "Đăng ký tài khoản", "POST", "/api/v1/auth/register", JWT_KHONG, "Đăng ký thành công", "", "Success", "201"),
    ("Auth", "Đăng ký tài khoản", "POST", "/api/v1/auth/register", JWT_KHONG, "Email đã được đăng ký", "EMAIL_ALREADY_EXISTS", "Error", "409"),
    ("Auth", "Đăng nhập", "POST", "/api/v1/auth/login", JWT_KHONG, "Đăng nhập thành công", "", "Success", "200"),
    ("Auth", "Đăng nhập", "POST", "/api/v1/auth/login", JWT_KHONG, "Sai email/mật khẩu", "INVALID_CREDENTIALS", "Error", "401"),
    ("Auth", "Đăng nhập", "POST", "/api/v1/auth/login", JWT_KHONG, "Email chưa xác minh", "EMAIL_NOT_VERIFIED", "Error", "403"),
    ("Auth", "Đăng nhập", "POST", "/api/v1/auth/login", JWT_KHONG, "Tài khoản bị vô hiệu hóa", "ACCOUNT_DISABLED", "Error", "403"),
    ("Auth", "Làm mới token", "POST", "/api/v1/auth/refresh", JWT_KHONG, "Cấp token mới", "", "Success", "200"),
    ("Auth", "Làm mới token", "POST", "/api/v1/auth/refresh", JWT_KHONG, "Refresh token không hợp lệ", "INVALID_TOKEN", "Error", "401"),
    ("Auth", "Làm mới token", "POST", "/api/v1/auth/refresh", JWT_KHONG, "Refresh token hết hạn", "TOKEN_EXPIRED", "Error", "401"),
    ("Auth", "Làm mới token", "POST", "/api/v1/auth/refresh", JWT_KHONG, "User không tồn tại", "USER_NOT_FOUND", "Error", "404"),
    ("Auth", "Làm mới token", "POST", "/api/v1/auth/refresh", JWT_KHONG, "Email chưa xác minh", "EMAIL_NOT_VERIFIED", "Error", "403"),
    ("Auth", "Làm mới token", "POST", "/api/v1/auth/refresh", JWT_KHONG, "Tài khoản bị vô hiệu hóa", "ACCOUNT_DISABLED", "Error", "403"),
    ("Auth", "Đăng xuất", "POST", "/api/v1/auth/logout", JWT_KHONG, "Thu hồi refresh token", "", "Success", "200"),
    ("Auth", "Xác minh email", "POST", "/api/v1/auth/verify-email", JWT_KHONG, "Xác minh thành công", "", "Success", "200"),
    ("Auth", "Xác minh email", "POST", "/api/v1/auth/verify-email", JWT_KHONG, "Token/OTP không hợp lệ", "INVALID_OTP", "Error", "400"),
    ("Auth", "Xác minh email", "POST", "/api/v1/auth/verify-email", JWT_KHONG, "Liên kết hết hạn", "OTP_EXPIRED", "Error", "400"),
    ("Auth", "Gửi lại email xác minh", "POST", "/api/v1/auth/resend-verification", JWT_KHONG, "Phản hồi chung (không lộ email có tồn tại hay không)", "", "Success", "200"),
    ("Auth", "Quên mật khẩu", "POST", "/api/v1/auth/forgot-password", JWT_KHONG, "Gửi mail nếu email tồn tại", "", "Success", "200"),
    ("Auth", "Đặt lại mật khẩu", "POST", "/api/v1/auth/reset-password", JWT_KHONG, "Đặt lại thành công", "", "Success", "200"),
    ("Auth", "Đặt lại mật khẩu", "POST", "/api/v1/auth/reset-password", JWT_KHONG, "Token không hợp lệ/đã dùng", "INVALID_OTP", "Error", "400"),
    ("Auth", "Đặt lại mật khẩu", "POST", "/api/v1/auth/reset-password", JWT_KHONG, "Token hết hạn", "OTP_EXPIRED", "Error", "400"),
    ("Auth", "Đặt lại mật khẩu", "POST", "/api/v1/auth/reset-password", JWT_KHONG, "User không tồn tại", "USER_NOT_FOUND", "Error", "404"),
    # --- User (me) ---
    ("User", "Lấy hồ sơ của tôi", "GET", "/api/v1/users/me/profile", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Cập nhật hồ sơ", "PUT", "/api/v1/users/me/profile", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Danh sách học vấn", "GET", "/api/v1/users/me/educations", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Thêm học vấn", "POST", "/api/v1/users/me/educations", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Cập nhật học vấn", "PUT", "/api/v1/users/me/educations/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Cập nhật học vấn", "PUT", "/api/v1/users/me/educations/{id}", JWT_CO, "Không tìm thấy bản ghi", "EDUCATION_NOT_FOUND", "Error", "404"),
    ("User", "Xóa học vấn", "DELETE", "/api/v1/users/me/educations/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Xóa học vấn", "DELETE", "/api/v1/users/me/educations/{id}", JWT_CO, "Không tìm thấy bản ghi", "EDUCATION_NOT_FOUND", "Error", "404"),
    ("User", "Danh sách ngôn ngữ", "GET", "/api/v1/users/me/languages", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Thêm ngôn ngữ", "POST", "/api/v1/users/me/languages", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Cập nhật ngôn ngữ", "PUT", "/api/v1/users/me/languages/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Cập nhật ngôn ngữ", "PUT", "/api/v1/users/me/languages/{id}", JWT_CO, "Không tìm thấy", "LANGUAGE_NOT_FOUND", "Error", "404"),
    ("User", "Xóa ngôn ngữ", "DELETE", "/api/v1/users/me/languages/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Xóa ngôn ngữ", "DELETE", "/api/v1/users/me/languages/{id}", JWT_CO, "Không tìm thấy", "LANGUAGE_NOT_FOUND", "Error", "404"),
    ("User", "Danh sách kinh nghiệm", "GET", "/api/v1/users/me/experiences", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Thêm kinh nghiệm", "POST", "/api/v1/users/me/experiences", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Cập nhật kinh nghiệm", "PUT", "/api/v1/users/me/experiences/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Cập nhật kinh nghiệm", "PUT", "/api/v1/users/me/experiences/{id}", JWT_CO, "Không tìm thấy", "EXPERIENCE_NOT_FOUND", "Error", "404"),
    ("User", "Xóa kinh nghiệm", "DELETE", "/api/v1/users/me/experiences/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Xóa kinh nghiệm", "DELETE", "/api/v1/users/me/experiences/{id}", JWT_CO, "Không tìm thấy", "EXPERIENCE_NOT_FOUND", "Error", "404"),
    ("User", "Danh sách chứng chỉ", "GET", "/api/v1/users/me/certificates", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Thêm chứng chỉ", "POST", "/api/v1/users/me/certificates", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Cập nhật chứng chỉ", "PUT", "/api/v1/users/me/certificates/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Cập nhật chứng chỉ", "PUT", "/api/v1/users/me/certificates/{id}", JWT_CO, "Không tìm thấy", "CERTIFICATE_NOT_FOUND", "Error", "404"),
    ("User", "Xóa chứng chỉ", "DELETE", "/api/v1/users/me/certificates/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("User", "Xóa chứng chỉ", "DELETE", "/api/v1/users/me/certificates/{id}", JWT_CO, "Không tìm thấy", "CERTIFICATE_NOT_FOUND", "Error", "404"),
    ("User", "Hồ sơ công khai theo user id", "GET", "/api/v1/users/{id}/profile", JWT_CO, "Thành công", "", "Success", "200"),
    # --- Mentor & packages ---
    ("Mentor", "Danh sách mentor đã xác minh", "GET", "/api/v1/mentors", JWT_CO, "Thành công", "", "Success", "200"),
    ("Mentor", "Chi tiết mentor", "GET", "/api/v1/mentors/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("Mentor", "Gói dịch vụ của mentor", "GET", "/api/v1/mentors/{id}/packages", JWT_CO, "Thành công", "", "Success", "200"),
    ("Mentor", "Cập nhật hồ sơ mentor (me)", "PUT", "/api/v1/mentors/me", JWT_CO, "Thành công", "", "Success", "200"),
    ("Mentor", "Tạo gói dịch vụ", "POST", "/api/v1/mentors/me/packages", JWT_CO, "Thành công", "", "Success", "200"),
    ("Mentor", "Xóa gói dịch vụ", "DELETE", "/api/v1/mentors/me/packages/{pkgId}", JWT_CO, "Thành công", "", "Success", "200"),
    ("Mentor", "Thêm mục curriculum", "POST", "/api/v1/mentors/me/packages/{pkgId}/versions/{verId}/curriculums", JWT_CO, "Thành công", "", "Success", "200"),
    ("Mentor", "Thêm mục curriculum", "POST", "/api/v1/mentors/me/packages/{pkgId}/versions/{verId}/curriculums", JWT_CO, "Phiên bản/gói không hợp lệ", "SERVICE_VERSION_NOT_FOUND", "Error", "404"),
    ("Mentor", "Thêm mục curriculum", "POST", "/api/v1/mentors/me/packages/{pkgId}/versions/{verId}/curriculums", JWT_CO, "Gói không active", "PACKAGE_INACTIVE", "Error", "400"),
    ("Mentor", "Liệt kê curriculum", "GET", "/api/v1/mentors/me/packages/{pkgId}/versions/{verId}/curriculums", JWT_CO, "Thành công", "", "Success", "200"),
    ("Mentor", "Liệt kê curriculum", "GET", "/api/v1/mentors/me/packages/{pkgId}/versions/{verId}/curriculums", JWT_CO, "Không tìm thấy phiên bản", "SERVICE_VERSION_NOT_FOUND", "Error", "404"),
    ("Mentor", "Xóa mục curriculum", "DELETE", "/api/v1/mentors/me/curriculums/{curriculumId}", JWT_CO, "Thành công", "", "Success", "200"),
    # --- Progress reports ---
    ("Báo cáo tiến độ", "Nộp báo cáo (mentee)", "POST", "/api/v1/mentee/reports", JWT_CO, "Thành công", "", "Success", "200"),
    ("Báo cáo tiến độ", "Danh sách báo cáo của tôi (mentee)", "GET", "/api/v1/mentee/reports", JWT_CO, "Thành công", "", "Success", "200"),
    ("Báo cáo tiến độ", "Báo cáo gán cho mentor", "GET", "/api/v1/mentors/me/reports", JWT_CO, "Thành công", "", "Success", "200"),
    ("Báo cáo tiến độ", "Phản hồi báo cáo (mentor)", "PUT", "/api/v1/mentors/me/reports/{id}/feedback", JWT_CO, "Thành công", "", "Success", "200"),
    # --- Orders ---
    ("Order", "Checkout — tạo đơn & URL VNPay", "POST", "/api/v1/orders/checkout", JWT_CO, "Tạo đơn thành công", "", "Success", "200"),
    ("Order", "Checkout — tạo đơn & URL VNPay", "POST", "/api/v1/orders/checkout", JWT_CO, "Phiên bản gói không tồn tại", "SERVICE_VERSION_NOT_FOUND", "Error", "404"),
    ("Order", "Checkout — tạo đơn & URL VNPay", "POST", "/api/v1/orders/checkout", JWT_CO, "Gói không active", "PACKAGE_INACTIVE", "Error", "400"),
    ("Order", "Danh sách đơn của tôi", "GET", "/api/v1/orders/me", JWT_CO, "Thành công", "", "Success", "200"),
    ("Order", "Chi tiết đơn", "GET", "/api/v1/orders/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("Order", "Chi tiết đơn", "GET", "/api/v1/orders/{id}", JWT_CO, "Không tìm thấy / không phải đơn của user", "ORDER_NOT_FOUND", "Error", "404"),
    # --- Payments ---
    # IPN: controller catches all exceptions → always HTTP 200; body is VNPay protocol string (not JSON ApiResponse).
    ("Payment", "VNPay IPN (server-to-server)", "GET", "/api/v1/payments/vnpay/ipn", JWT_KHONG, "Xử lý callback OK (body: 00|Confirm Success)", "", "Success", "200"),
    ("Payment", "VNPay IPN (server-to-server)", "GET", "/api/v1/payments/vnpay/ipn", JWT_KHONG, "Callback thất bại hoặc exception (body: 01|Error; không trả JSON errors.type)", "(xem log server)", "Error", "200"),
    ("Payment", "VNPay return (redirect browser)", "GET", "/api/v1/payments/vnpay/return", JWT_KHONG, "JSON ApiResponse (có thể success hoặc failed trong message)", "", "Success", "200"),
    ("Payment", "Payment theo orderId", "GET", "/api/v1/payments/order/{orderId}", JWT_CO, "Thành công", "", "Success", "200"),
    ("Payment", "Payment theo orderId", "GET", "/api/v1/payments/order/{orderId}", JWT_CO, "Không tìm thấy", "PAYMENT_NOT_FOUND", "Error", "404"),
    # --- Bookings ---
    ("Booking", "Booking của tôi (buyer)", "GET", "/api/v1/bookings/me/buyer", JWT_CO, "Thành công", "", "Success", "200"),
    ("Booking", "Booking của tôi (mentor)", "GET", "/api/v1/bookings/me/mentor", JWT_CO, "Thành công", "", "Success", "200"),
    ("Booking", "Chi tiết booking", "GET", "/api/v1/bookings/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("Booking", "Chi tiết booking", "GET", "/api/v1/bookings/{id}", JWT_CO, "Không tìm thấy", "BOOKING_NOT_FOUND", "Error", "404"),
    ("Booking", "Chi tiết booking", "GET", "/api/v1/bookings/{id}", JWT_CO, "Không có quyền xem", "BOOKING_ACCESS_DENIED", "Error", "403"),
    ("Booking", "Cập nhật phiên (session)", "PATCH", "/api/v1/bookings/{bookingId}/sessions/{sessionId}", JWT_CO, "Thành công", "", "Success", "200"),
    ("Booking", "Cập nhật phiên (session)", "PATCH", "/api/v1/bookings/{bookingId}/sessions/{sessionId}", JWT_CO, "Booking/session không tìm thấy hoặc không có quyền", "BOOKING_NOT_FOUND", "Error", "404"),
    ("Booking", "Cập nhật phiên (session)", "PATCH", "/api/v1/bookings/{bookingId}/sessions/{sessionId}", JWT_CO, "Không có quyền", "BOOKING_ACCESS_DENIED", "Error", "403"),
    ("Booking", "Cập nhật phiên (session)", "PATCH", "/api/v1/bookings/{bookingId}/sessions/{sessionId}", JWT_CO, "Session không thuộc booking", "SESSION_NOT_FOUND", "Error", "404"),
    ("Booking", "Thêm minh chứng buổi học", "POST", "/api/v1/bookings/{bookingId}/sessions/{sessionId}/evidences", JWT_CO, "Thành công", "", "Success", "200"),
    ("Booking", "Thêm minh chứng buổi học", "POST", "/api/v1/bookings/{bookingId}/sessions/{sessionId}/evidences", JWT_CO, "Lỗi tương tự cập nhật session", "BOOKING_NOT_FOUND", "Error", "404"),
    ("Booking", "Thêm minh chứng buổi học", "POST", "/api/v1/bookings/{bookingId}/sessions/{sessionId}/evidences", JWT_CO, "Không có quyền", "BOOKING_ACCESS_DENIED", "Error", "403"),
    # --- Chat ---
    ("Chat", "Tạo conversation", "POST", "/api/v1/chat/conversations", JWT_CO, "Thành công", "", "Success", "200"),
    ("Chat", "Danh sách conversation", "GET", "/api/v1/chat/conversations", JWT_CO, "Thành công", "", "Success", "200"),
    ("Chat", "Tin nhắn trong conversation", "GET", "/api/v1/chat/conversations/{conversationId}/messages", JWT_CO, "Thành công", "", "Success", "200"),
    ("Chat", "Tin nhắn trong conversation", "GET", "/api/v1/chat/conversations/{conversationId}/messages", JWT_CO, "Không tìm thấy", "CONVERSATION_NOT_FOUND", "Error", "404"),
    ("Chat", "Tin nhắn trong conversation", "GET", "/api/v1/chat/conversations/{conversationId}/messages", JWT_CO, "Không có quyền", "CHAT_ACCESS_DENIED", "Error", "403"),
    ("Chat", "Gửi tin nhắn", "POST", "/api/v1/chat/conversations/{conversationId}/messages", JWT_CO, "Thành công", "", "Success", "200"),
    ("Chat", "Gửi tin nhắn", "POST", "/api/v1/chat/conversations/{conversationId}/messages", JWT_CO, "Không tìm thấy / không có quyền", "CONVERSATION_NOT_FOUND", "Error", "404"),
    ("Chat", "Gửi tin nhắn", "POST", "/api/v1/chat/conversations/{conversationId}/messages", JWT_CO, "Không có quyền", "CHAT_ACCESS_DENIED", "Error", "403"),
    # --- Files ---
    ("File", "Upload file", "POST", "/api/v1/files", JWT_CO, "Thành công", "", "Success", "200"),
    ("File", "Metadata file", "GET", "/api/v1/files/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("File", "Metadata file", "GET", "/api/v1/files/{id}", JWT_CO, "Không tìm thấy", "FILE_NOT_FOUND", "Error", "404"),
    ("File", "Metadata file", "GET", "/api/v1/files/{id}", JWT_CO, "Không có quyền", "FILE_ACCESS_DENIED", "Error", "403"),
    ("File", "Xóa mềm file", "DELETE", "/api/v1/files/{id}", JWT_CO, "Thành công", "", "Success", "200"),
    ("File", "Xóa mềm file", "DELETE", "/api/v1/files/{id}", JWT_CO, "Không tìm thấy / không có quyền", "FILE_NOT_FOUND", "Error", "404"),
    # --- Trust ---
    ("Trust", "Tạo báo cáo kiểm duyệt", "POST", "/api/v1/trust/reports", JWT_CO, "Thành công", "", "Success", "200"),
    ("Trust", "Báo cáo của tôi", "GET", "/api/v1/trust/reports/me", JWT_CO, "Thành công", "", "Success", "200"),
    ("Trust", "Thêm bằng chứng báo cáo", "POST", "/api/v1/trust/reports/{reportId}/evidences", JWT_CO, "Thành công", "", "Success", "200"),
    ("Trust", "Thêm bằng chứng báo cáo", "POST", "/api/v1/trust/reports/{reportId}/evidences", JWT_CO, "Không tìm thấy báo cáo", "REPORT_NOT_FOUND", "Error", "404"),
    ("Trust", "Thêm bằng chứng báo cáo", "POST", "/api/v1/trust/reports/{reportId}/evidences", JWT_CO, "Không có quyền", "TRUST_ACCESS_DENIED", "Error", "403"),
    ("Trust", "Giải quyết báo cáo (mod)", "PUT", "/api/v1/trust/reports/{reportId}/resolve", JWT_CO, "Thành công", "", "Success", "200"),
    ("Trust", "Giải quyết báo cáo (mod)", "PUT", "/api/v1/trust/reports/{reportId}/resolve", JWT_CO, "Không tìm thấy", "REPORT_NOT_FOUND", "Error", "404"),
    ("Trust", "Tạo tranh chấp", "POST", "/api/v1/trust/disputes", JWT_CO, "Thành công", "", "Success", "200"),
    ("Trust", "Tranh chấp của tôi", "GET", "/api/v1/trust/disputes/me", JWT_CO, "Thành công", "", "Success", "200"),
    ("Trust", "Giải quyết tranh chấp (mod)", "PUT", "/api/v1/trust/disputes/{disputeId}/resolve", JWT_CO, "Thành công", "", "Success", "200"),
    ("Trust", "Giải quyết tranh chấp (mod)", "PUT", "/api/v1/trust/disputes/{disputeId}/resolve", JWT_CO, "Không tìm thấy", "DISPUTE_NOT_FOUND", "Error", "404"),
]

ERROR_CODE_TABLE = [
    ("AuthErrorCode", "EMAIL_ALREADY_EXISTS", "409"),
    ("AuthErrorCode", "INVALID_CREDENTIALS", "401"),
    ("AuthErrorCode", "EMAIL_NOT_VERIFIED", "403"),
    ("AuthErrorCode", "ACCOUNT_DISABLED", "403"),
    ("AuthErrorCode", "INVALID_TOKEN", "401"),
    ("AuthErrorCode", "TOKEN_EXPIRED", "401"),
    ("AuthErrorCode", "INVALID_OTP", "400"),
    ("AuthErrorCode", "OTP_EXPIRED", "400"),
    ("AuthErrorCode", "OTP_ALREADY_USED", "400"),
    ("AuthErrorCode", "USER_NOT_FOUND", "404"),
    ("AuthErrorCode", "ACCESS_DENIED", "403"),
    ("UserErrorCode", "PROFILE_NOT_FOUND", "404"),
    ("UserErrorCode", "EDUCATION_NOT_FOUND", "404"),
    ("UserErrorCode", "LANGUAGE_NOT_FOUND", "404"),
    ("UserErrorCode", "EXPERIENCE_NOT_FOUND", "404"),
    ("UserErrorCode", "CERTIFICATE_NOT_FOUND", "404"),
    ("OrderErrorCode", "ORDER_NOT_FOUND", "404"),
    ("OrderErrorCode", "PAYMENT_FAILED", "402"),
    ("OrderErrorCode", "PAYMENT_INVALID_SIGNATURE", "400"),
    ("PaymentErrorCode", "PAYMENT_NOT_FOUND", "404"),
    ("PaymentErrorCode", "PAYMENT_FAILED", "402"),
    ("PaymentErrorCode", "INVALID_SIGNATURE", "400"),
    ("PaymentErrorCode", "PAYMENT_ALREADY_PROCESSED", "409"),
    ("PaymentErrorCode", "SIGNATURE_COMPUTATION_FAILED", "500"),
    ("BookingErrorCode", "BOOKING_NOT_FOUND", "404"),
    ("BookingErrorCode", "BOOKING_ACCESS_DENIED", "403"),
    ("BookingErrorCode", "SESSION_NOT_FOUND", "404"),
    ("ChatErrorCode", "CONVERSATION_NOT_FOUND", "404"),
    ("ChatErrorCode", "CHAT_ACCESS_DENIED", "403"),
    ("FileErrorCode", "FILE_NOT_FOUND", "404"),
    ("FileErrorCode", "FILE_ACCESS_DENIED", "403"),
    ("FileErrorCode", "FILE_UPLOAD_FAILED", "500"),
    ("ServiceErrorCode", "SERVICE_VERSION_NOT_FOUND", "404"),
    ("ServiceErrorCode", "PACKAGE_INACTIVE", "400"),
    ("ServiceErrorCode", "MENTOR_NOT_FOUND", "404"),
    ("ServiceErrorCode", "PACKAGE_NOT_FOUND", "404"),
    ("ServiceErrorCode", "CURRICULUM_NOT_FOUND", "404"),
    ("ProgressReportErrorCode", "PROGRESS_REPORT_NOT_FOUND", "404"),
    ("ProgressReportErrorCode", "PROGRESS_REPORT_ACCESS_DENIED", "403"),
    ("CommonErrorCode", "BAD_REQUEST", "400"),
    ("CommonErrorCode", "VALIDATION_ERROR", "400"),
    ("CommonErrorCode", "UNAUTHORIZED", "401"),
    ("CommonErrorCode", "FORBIDDEN", "403"),
    ("CommonErrorCode", "PAYLOAD_TOO_LARGE", "413"),
    ("CommonErrorCode", "INTERNAL_ERROR", "500"),
    ("TrustErrorCode", "REPORT_NOT_FOUND", "404"),
    ("TrustErrorCode", "DISPUTE_NOT_FOUND", "404"),
    ("TrustErrorCode", "TRUST_ACCESS_DENIED", "403"),
]


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws):
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        maxlen = len(HEADERS[col - 1])
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                maxlen = max(maxlen, min(len(str(v)), 80))
        ws.column_dimensions[letter].width = maxlen + 2


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Sheet 1: catalog ---
    ws1 = wb.active
    ws1.title = "Danh mục API"
    ws1.append(HEADERS)
    style_header(ws1)
    for r in ROWS:
        ws1.append(list(r))
    for row in range(2, ws1.max_row + 1):
        for col in range(1, len(HEADERS) + 1):
            ws1.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws1)

    # --- Sheet 2: error codes ---
    ws2 = wb.create_sheet("Bảng mã lỗi (enum)")
    h2 = ["Enum Java", "errors.type", "HTTP mặc định"]
    ws2.append(h2)
    for c in range(1, 4):
        cell = ws2.cell(row=1, column=c)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF")
    for row in ERROR_CODE_TABLE:
        ws2.append(list(row))
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 16

    # --- Sheet 3: guide + future ---
    ws3 = wb.create_sheet("Hướng dẫn & mở rộng")
    notes = [
        "UniShare API — catalog được sinh từ source (controllers + *ErrorCode).",
        "",
        "Cột « JWT »: Không = public theo SecurityConfig; Có = cần Bearer JWT.",
        "Lỗi nghiệp vụ: body ApiResponse với errors.type = mã trong cột « Mã lỗi ».",
        "401 Unauthorized: thiếu/sai JWT (Spring Security), không qua AppException.",
        "403 Forbidden: không đủ quyền (ví dụ @PreAuthorize) hoặc mã ACCESS_DENIED.",
        "400 Bad Request: validation @Valid hoặc mã 4xx tương ứng trong enum.",
        "",
        "Đăng nhập thành công trả HTTP 200 (không phải 201). Đăng ký thành công: 201.",
        "",
        "Hàng mẫu bên dưới — copy khi thêm API mới:",
    ]
    for i, line in enumerate(notes, start=1):
        ws3.cell(row=i, column=1, value=line)
    start = len(notes) + 2
    ws3.cell(row=start, column=1, value="Module")
    ws3.cell(row=start, column=2, value="Tên API (VN)")
    ws3.cell(row=start, column=3, value="Method")
    ws3.cell(row=start, column=4, value="Endpoint")
    ws3.cell(row=start, column=5, value="JWT")
    ws3.cell(row=start, column=6, value="Kịch bản / mô tả")
    ws3.cell(row=start, column=7, value="Mã lỗi (errors.type)")
    ws3.cell(row=start, column=8, value="Kết quả")
    ws3.cell(row=start, column=9, value="HTTP")
    for r in range(start, start + 8):
        for c in range(1, 10):
            ws3.cell(row=r, column=c).fill = PatternFill("solid", fgColor="E7E6E6")

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
